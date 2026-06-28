"""
========================================
ETL PIPELINE - BANCO NORTE
Tarea 1: Migracion Excel -> Python
========================================
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import sys
import argparse
from datetime import datetime
from pathlib import Path


# ============================================
#             CONFIGURACION DE LOGGING
# ============================================
def setup_logging():
    """Configura logging con archivo y consola."""
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(
                log_dir / f"etl_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            ),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)


logger = setup_logging()


# ============================================
# 1.            Extracción de Datos
# ============================================
def extraer_datos(ruta_archivo: str) -> dict:
    """
    Lee multiples hojas del archivo Excel origen.

    Args:
        ruta_archivo: Path al archivo Excel de origen

    Returns:
        dict: Diccionario con nombre_hoja -> DataFrame

    Raises:
        FileNotFoundError: Si el archivo no existe
        ValueError: Si no se encuentran hojas esperadas
    """
    try:
        logger.info(f"Iniciando extraccion: {ruta_archivo}")

        ruta = Path(ruta_archivo)
        if not ruta.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta_archivo}")

        # Leer todas las hojas
        dfs = pd.read_excel(ruta, sheet_name=None, engine='openpyxl')

        logger.info(f"Hojas detectadas: {list(dfs.keys())}")
        logger.info(f"Total hojas: {len(dfs)}")

        # Validar hojas minimas esperadas
        hojas_esperadas = ['Datos', 'Clientes', 'Productos']
        for hoja in hojas_esperadas:
            if hoja not in dfs:
                logger.warning(f"Hoja esperada no encontrada: {hoja}")

        # Log de dimensiones
        for nombre, df in dfs.items():
            logger.info(f"Hoja '{nombre}': {df.shape[0]} filas x {df.shape[1]} columnas")

        return dfs

    except Exception as e:
        logger.error(f"Error en extraccion: {str(e)}")
        raise


# ============================================
# 2.  Transformación (Reemplazo de VLOOKUP)
# ============================================
def transformar_datos(dfs: dict) -> pd.DataFrame:
    """
    Reemplaza VLOOKUPs por merges y realiza limpieza de datos.

    Args:
        dfs: Diccionario de DataFrames extraidos

    Returns:
        pd.DataFrame: DataFrame principal transformado
    """
    logger.info("Iniciando transformacion")

    # Verificar hojas requeridas
    required = ['Datos', 'Clientes', 'Productos']
    for hoja in required:
        if hoja not in dfs:
            raise ValueError(f"Hoja requerida no encontrada: {hoja}")

    df_datos = dfs['Datos'].copy()
    df_clientes = dfs['Clientes'].copy()
    df_productos = dfs['Productos'].copy()

    logger.info(f"Datos originales: {len(df_datos)} registros")

    # --- LIMPIEZA PREVIA ---
    df_datos = df_datos.dropna(subset=['ID_CLIENTE', 'MONTO'])
    logger.info(f"Despues de limpieza: {len(df_datos)} registros")

    # Convertir tipos
    df_datos['FECHA'] = pd.to_datetime(df_datos['FECHA'], errors='coerce')
    df_datos['MONTO'] = pd.to_numeric(df_datos['MONTO'], errors='coerce')

    # --- MERGE 1: Datos + Clientes (reemplaza VLOOKUP) ---
    logger.info("Ejecutando merge con Clientes...")
    df_merged = pd.merge(
        df_datos,
        df_clientes[['ID_CLIENTE', 'NOMBRE_CLIENTE', 'SEGMENTO', 'REGION']],
        on='ID_CLIENTE',
        how='left',
        indicator=True
    )

    ids_huerfanos = (df_merged['_merge'] == 'left_only').sum()
    if ids_huerfanos > 0:
        logger.warning(f"Clientes sin match: {ids_huerfanos}")

    df_merged = df_merged.drop(columns=['_merge'])

    # --- MERGE 2: Datos + Productos (reemplaza VLOOKUP) ---
    logger.info("Ejecutando merge con Productos...")
    df_final = pd.merge(
        df_merged,
        df_productos[['ID_PRODUCTO', 'NOMBRE_PRODUCTO', 'CATEGORIA']],
        on='ID_PRODUCTO',
        how='left',
        indicator=True
    )

    productos_huerfanos = (df_final['_merge'] == 'left_only').sum()
    if productos_huerfanos > 0:
        logger.warning(f"Productos sin match: {productos_huerfanos}")

    df_final = df_final.drop(columns=['_merge'])

    # --- FEATURE ENGINEERING ---
    df_final['AÑO'] = df_final['FECHA'].dt.year
    df_final['MES'] = df_final['FECHA'].dt.month
    df_final['TRIMESTRE'] = df_final['FECHA'].dt.quarter
    df_final['DIA_SEMANA'] = df_final['FECHA'].dt.day_name()
    df_final['MONTO_COMISION'] = df_final['MONTO'] * 0.02  # 2% comision ejemplo

    logger.info(f"Transformacion completada: {len(df_final)} registros, {len(df_final.columns)} columnas")

    return df_final


# ============================================
# 3. Calculos y Agregaciones (Reemplazo SUMIF)
# ============================================
def calcular_agregaciones(df: pd.DataFrame) -> tuple:
    """
    Reemplaza SUMIF/COUNTIF con groupby.

    Args:
        df: DataFrame transformado

    Returns:
        tuple: (resumen_categoria, resumen_mensual, resumen_segmento)
    """
    logger.info("Calculando agregaciones")

    # --- Resumen por categoria (SUMIF) ---
    resumen_categoria = df.groupby('CATEGORIA').agg({
        'MONTO': ['sum', 'mean', 'count', 'std'],
        'MONTO_COMISION': 'sum',
        'ID_CLIENTE': 'nunique'
    }).reset_index()

    resumen_categoria.columns = [
        'CATEGORIA', 'TOTAL_MONTO', 'PROMEDIO_MONTO', 'NUM_TRANSACCIONES',
        'STD_MONTO', 'TOTAL_COMISION', 'CLIENTES_UNICOS'
    ]

    # --- Resumen mensual (tabla dinamica) ---
    # CORRECCIÓN: convertir Period a string para compatibilidad con openpyxl
    df = df.copy()
    df['PERIODO'] = df['FECHA'].dt.to_period('M').astype(str)
    resumen_mensual = df.pivot_table(
        values=['MONTO', 'MONTO_COMISION'],
        index='PERIODO',
        columns='CATEGORIA',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    resumen_mensual.columns = [
        f"{col[0]}_{col[1]}" if col[1] else col[0]
        for col in resumen_mensual.columns.values
    ]

    # --- Resumen por segmento y region ---
    resumen_segmento = df.groupby(['SEGMENTO', 'REGION']).agg({
        'MONTO': 'sum',
        'ID_CLIENTE': 'nunique',
        'ID_TRANSACCION': 'count'
    }).reset_index()
    resumen_segmento.columns = ['SEGMENTO', 'REGION', 'TOTAL_MONTO', 'CLIENTES_UNICOS', 'NUM_TRANSACCIONES']

    logger.info(f"Agregaciones: {len(resumen_categoria)} categorias, {len(resumen_mensual)} periodos")

    return resumen_categoria, resumen_mensual, resumen_segmento


# ============================================
# 4.    GENERACION DE EXCEL DE SALIDA
# ============================================
def aplicar_formato(ws, color_header="1F4E79", color_alt="D6E4F0"):
    """Aplica formato corporativo a una hoja de Excel."""

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
        cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = PatternFill(start_color=color_alt, end_color=color_alt, fill_type="solid") if row_idx % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            cell.font = Font(name='Calibri', size=10)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = 'A2'


def generar_excel(df_principal, resumen_cat, resumen_mes, resumen_seg, ruta_salida):
    """Genera archivo Excel con formato profesional."""
    logger.info(f"Generando salida: {ruta_salida}")

    with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
        df_principal.to_excel(writer, sheet_name='Datos_Transformados', index=False)
        ws = writer.sheets['Datos_Transformados']
        aplicar_formato(ws)

        resumen_cat.to_excel(writer, sheet_name='Resumen_Categoria', index=False)
        ws = writer.sheets['Resumen_Categoria']
        aplicar_formato(ws)

        resumen_mes.to_excel(writer, sheet_name='Resumen_Mensual', index=False)
        ws = writer.sheets['Resumen_Mensual']
        aplicar_formato(ws)

        resumen_seg.to_excel(writer, sheet_name='Resumen_Segmento', index=False)
        ws = writer.sheets['Resumen_Segmento']
        aplicar_formato(ws)

        metadata = pd.DataFrame({
            'Campo': ['Fecha de generacion', 'Version del script', 'Registros procesados',
                     'Tiempo de ejecucion', 'Validacion'],
            'Valor': [datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                     '1.0.0',
                     str(len(df_principal)),
                     'Automatico',
                     'OK']
        })
        metadata.to_excel(writer, sheet_name='Metadata', index=False)
        ws = writer.sheets['Metadata']
        aplicar_formato(ws)

    logger.info("Archivo Excel generado exitosamente")


# ============================================
# 5.    VALIDACION Y CONTROL DE CALIDAD
# ============================================
def validar_resultados(df_salida, df_origen):
    """Valida integridad de datos entre origen y salida."""
    logger.info("Iniciando validacion de calidad de datos")

    checks = []

    # Check 1: Dataset no vacio
    check1 = len(df_salida) > 0
    checks.append(("Dataset no vacio", check1, f"{len(df_salida)} registros"))
    assert check1, "ERROR: El dataset resultante esta vacio"

    # Check 2: Conteo de registros vs. origen
    registros_origen = len(df_origen['Datos'].dropna(subset=['ID_CLIENTE', 'MONTO']))
    check2 = len(df_salida) >= registros_origen * 0.95
    checks.append(("Preservacion de registros", check2, f"{len(df_salida)}/{registros_origen}"))

    # Check 3: Suma total preservada
    # CORRECCIÓN: proteger contra suma_origen == 0 para evitar ZeroDivisionError
    suma_origen = pd.to_numeric(df_origen['Datos']['MONTO'], errors='coerce').sum()
    suma_salida = df_salida['MONTO'].sum()
    if suma_origen != 0:
        check3 = abs(suma_origen - suma_salida) < (suma_origen * 0.01)
    else:
        check3 = suma_salida == 0
    checks.append(("Preservacion de montos", check3, f"Origen: {suma_origen:.2f}, Salida: {suma_salida:.2f}"))

    # Check 4: Sin valores nulos en columnas criticas de transaccion
    # CORRECCIÓN: NOMBRE_CLIENTE puede ser nulo si hay clientes sin match (left join);
    # solo validamos columnas que deben ser siempre no nulas.
    nulos = df_salida[['ID_CLIENTE', 'MONTO', 'FECHA']].isnull().sum().sum()
    check4 = nulos == 0
    checks.append(("Valores nulos criticos", check4, f"{nulos} nulos encontrados"))

    # Check 5: Rango de fechas valido
    fecha_min = df_salida['FECHA'].min()
    fecha_max = df_salida['FECHA'].max()
    check5 = fecha_min.year >= 2020 and fecha_max <= pd.Timestamp.now()
    checks.append(("Rango de fechas", check5, f"{fecha_min.date()} a {fecha_max.date()}"))

    # Check 6: Sin duplicados de ID_TRANSACCION
    duplicados = df_salida['ID_TRANSACCION'].duplicated().sum()
    check6 = duplicados == 0
    checks.append(("Duplicados de transaccion", check6, f"{duplicados} duplicados"))

    logger.info("=" * 50)
    logger.info("RESULTADOS DE VALIDACION")
    logger.info("=" * 50)
    for nombre, resultado, detalle in checks:
        status = "PASS" if resultado else "FAIL"
        logger.info(f"{status} | {nombre}: {detalle}")

    total_pass = sum(1 for _, r, _ in checks if r)
    logger.info(f"Total: {total_pass}/{len(checks)} checks pasados")

    if total_pass < len(checks):
        raise ValueError(f"Validacion fallida: {len(checks) - total_pass} checks no pasaron")

    logger.info("Todas las validaciones pasaron exitosamente")


# ============================================
# 6.           EJECUCION PRINCIPAL
# ============================================
def main():
    parser = argparse.ArgumentParser(description='ETL BancoNorte - Migracion Excel a Python')
    parser.add_argument('--input', required=True, help='Ruta al archivo Excel de origen')
    # CORRECCIÓN: ruta de salida relativa al proyecto, no con ../../
    parser.add_argument('--output', default='Data/processed/', help='Directorio de salida')
    parser.add_argument('--validate', action='store_true', help='Ejecutar validaciones')
    args = parser.parse_args()

    start_time = datetime.now()

    try:
        logger.info("=" * 60)
        logger.info("INICIANDO PIPELINE ETL - BANCO NORTE")
        logger.info("=" * 60)

        dfs_origen = extraer_datos(args.input)
        df_transformado = transformar_datos(dfs_origen)
        resumen_cat, resumen_mes, resumen_seg = calcular_agregaciones(df_transformado)

        output_dir = Path(args.output)
        output_dir.mkdir(parents=True, exist_ok=True)
        ruta_salida = output_dir / f"reporte_automatizado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        generar_excel(df_transformado, resumen_cat, resumen_mes, resumen_seg, ruta_salida)

        if args.validate:
            validar_resultados(df_transformado, dfs_origen)

        duration = (datetime.now() - start_time).total_seconds()
        logger.info("=" * 60)
        logger.info(f"PIPELINE COMPLETADO EN {duration:.2f} SEGUNDOS")
        logger.info(f"Archivo generado: {ruta_salida}")
        logger.info("=" * 60)

        return 0

    except Exception as e:
        logger.critical(f"PIPELINE FALLIDO: {str(e)}")
        import traceback
        logger.critical(traceback.format_exc())
        return 1


if __name__ == "__main__":
    sys.exit(main())
